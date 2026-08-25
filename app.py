import io
import os
import re
import time
import xml.etree.ElementTree as ET
import google.generativeai as genai
import openpyxl
import pandas as pd
import requests
import rispy  # GIE RIS 파일 파싱용
import streamlit as st

st.set_page_config(
    page_title="Taewoong Medical - AI 문헌 스크리닝", layout="wide"
)

# --------------------------------------------------
# 브라우저 새로고침(F5) 및 탭 닫기 이탈 방지 스크립트
# --------------------------------------------------
st.components.v1.html(
    """
    <script>
    window.addEventListener('beforeunload', function (e) {
        e.preventDefault();
        e.returnValue = '';
    });
    </script>
    """,
    height=0,
)

# --------------------------------------------------
# 마크다운 별표(**)를 유니코드 굵은 글씨로 변환하는 함수
# --------------------------------------------------
def to_unicode_bold(text):
    if not text:
        return text
    bold_map = {
        'A': '𝐀', 'B': '𝐁', 'C': '𝐂', 'D': '𝐃', 'E': '𝐄', 'F': '𝐅', 'G': '𝐆', 'H': '𝐇', 'I': '𝐈', 'J': '𝐉', 'K': '𝐊', 'L': '𝐋', 'M': '𝐌', 'N': '𝐍', 'O': '𝐎', 'P': '𝐏', 'Q': '𝐐', 'R': '𝐑', 'S': '𝐒', 'T': '𝐓', 'U': '𝐔', 'V': '𝐕', 'W': '𝐖', 'X': '𝐗', 'Y': '𝐘', 'Z': '𝐙',
        'a': '𝐚', 'b': '𝐛', 'c': '𝐜', 'd': '𝐝', 'e': '𝐞', 'f': '𝐟', 'g': '𝐠', 'h': '𝐡', 'i': '𝐢', 'j': '𝐣', 'k': '𝐤', 'l': '𝐥', 'm': '𝐦', 'n': '𝐧', 'o': '𝐨', 'p': '𝐩', 'q': '𝐪', 'r': '𝐫', 's': '𝐬', 't': '𝐭', 'u': '𝐮', 'v': '𝐯', 'w': '𝐰', 'x': '𝐱', 'y': '𝐲', 'z': '𝐳'
    }
    def replace_bold(match):
        word = match.group(1)
        return "".join(bold_map.get(c, c) for c in word)
    
    return re.sub(r'\*\*(.*?)\*\*', replace_bold, text)

# --------------------------------------------------
# Excel(.xlsx) 변환 헬퍼 함수
# --------------------------------------------------
def convert_df_to_excel(df_input):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    bold_reverse_map = {
        '𝐀': 'A', '𝐁': 'B', '𝐂': 'C', '𝐃': 'D', '𝐄': 'E', '𝐅': 'F', '𝐆': 'G', '𝐇': 'H', '𝐈': 'I', '𝐉': 'J', '𝐊': 'K', '𝐋': 'L', '𝐌': 'M', '𝐍': 'N', '𝐎': 'O', '𝐏': 'P', '𝐐': 'Q', '𝐑': 'R', '𝐒': 'S', '𝐓': 'T', '𝐔': 'U', '𝐕': 'V', '𝐖': 'W', '𝐗': 'X', '𝐘': 'Y', '𝐙': 'Z',
        '𝐚': 'a', '𝐛': 'b', '𝐜': 'c', 'd': 'd', '𝐞': 'e', '𝐟': 'f', '𝐠': 'g', '𝐡': 'h', '𝐢': 'i', '𝐣': 'j', '𝐤': 'k', '𝐥': 'l', '𝐦': 'm', '𝐧': 'n', '𝐨': 'o', '𝐩': 'p', '𝐪': 'q', '𝐫': 'r', '𝐬': 's', '𝐭': 't', '𝐮': 'u', '𝐯': 'v', '𝐰': 'w', '𝐱': 'x', '𝐲': 'y', '𝐳': 'z'
    }
    
    df_clean = df_input.copy()
    for col in df_clean.columns:
        if df_clean[col].dtype == 'object':
            df_clean[col] = df_clean[col].astype(str).apply(
                lambda text: "".join(bold_reverse_map.get(c, c) for c in text)
            )

    excel_io = io.BytesIO()
    
    with pd.ExcelWriter(excel_io, engine='openpyxl') as writer:
        df_clean.to_excel(writer, index=False, sheet_name='Screening Results')
        ws = writer.sheets['Screening Results']

        header_fill = PatternFill(start_color="0B1A2D", end_color="0B1A2D", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        body_font = Font(name="Calibri", size=10)
        
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = body_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                
                val_str = str(cell.value or "")
                if val_str.startswith("Include"):
                    cell.font = Font(name="Calibri", size=10, bold=True, color="166534")
                elif val_str.startswith("Exclude"):
                    cell.font = Font(name="Calibri", size=10, bold=True, color="991B1B")

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

    excel_io.seek(0)
    return excel_io.getvalue()

# --------------------------------------------------
# 🎨 고급 커스텀 CSS
# --------------------------------------------------
st.markdown(
    """
<style>
    .block-container {
        padding-top: 2rem !important;
        padding-bottom: 3rem !important;
        max-width: 96% !important;
    }
    html, body, [data-testid="stAppViewContainer"] {
        overflow-y: auto !important;
        height: auto !important;
    }
    section[data-testid="stSidebar"] {
        font-size: 15px !important;
    }
    section[data-testid="stSidebar"] label {
        font-size: 14px !important;
        font-weight: 700 !important;
    }
    section[data-testid="stSidebar"] input {
        font-size: 14px !important;
        padding: 8px 12px !important;
    }
    div[data-testid="stFileUploader"] label p {
        letter-spacing: -0.8px !important;
        white-space: nowrap !important;
    }
    
    .hero-container {
        background: rgba(15, 23, 42, 0.92);
        backdrop-filter: blur(8px);
        border: 1px solid rgba(255, 255, 255, 0.08);
        padding: 28px 32px;
        border-radius: 16px;
        color: #ffffff;
        box-shadow: 0 10px 25px -5px rgba(15, 23, 42, 0.15);
        margin-bottom: 20px;
    }
    .hero-header-flex { display: flex; justify-content: space-between; align-items: center; margin-bottom: 10px; }
    .hero-tag { background: linear-gradient(90deg, #84cc16 0%, #06b6d4 100%); color: #ffffff; font-size: 12px; font-weight: 800; padding: 4px 12px; border-radius: 20px; text-transform: uppercase; }
    .dept-tag { color: #94a3b8; font-size: 13px; font-weight: 600; }
    .hero-title { font-size: 26px; font-weight: 800; color: #ffffff; margin: 4px 0px 6px 0px; letter-spacing: -0.5px; }
    .hero-subtitle { font-size: 14px; color: #cbd5e1; margin-bottom: 0px; font-weight: 400; }
    .card-title { font-size: 12px; font-weight: 700; color: #0284c7; text-transform: uppercase; margin-bottom: 4px; }
    .card-value { font-size: 14px; font-weight: 700; color: #0f172a; margin-bottom: 4px; letter-spacing: -0.3px; white-space: normal !important; word-break: keep-all !important; }
    .card-desc { font-size: 12px; color: #64748b; margin-bottom: 12px; word-break: keep-all !important; }
    
    .selected-category-box { 
        background: rgba(15, 23, 42, 0.92);
        backdrop-filter: blur(8px);
        border: 1px solid rgba(255, 255, 255, 0.08);
        border-radius: 12px; 
        padding: 20px 26px; 
        margin-bottom: 20px; 
        box-shadow: 0 4px 20px -2px rgba(15, 23, 42, 0.12);
    }
    .selected-category-label { 
        font-size: 11px; 
        font-weight: 700; 
        color: #38bdf8; 
        text-transform: uppercase; 
        letter-spacing: 1.2px; 
        margin-bottom: 6px; 
    }
    .selected-category-title { 
        font-size: 22px; 
        font-weight: 800; 
        color: #ffffff; 
        letter-spacing: -0.5px;
        margin: 0; 
    }
    
    div[data-testid="stSegmentedControl"] { background-color: #f1f5f9; padding: 6px; border-radius: 12px; border: 1px solid #e2e8f0; display: flex !important; flex-wrap: nowrap !important; overflow-x: auto !important; }
    div[data-testid="stSegmentedControl"] button { border-radius: 8px !important; font-weight: 1000 !important; font-size: 13px !important; border: none !important; padding: 8px 14px !important; white-space: nowrap !important; transition: all 0.2s ease !important; flex: 1 1 auto !important; }
    div[data-testid="stSegmentedControl"] button[aria-selected="true"] { background-color: #0b1a2d !important; color: #ffffff !important; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1) !important; }
    .prod-item-title { font-weight: 700; font-size: 15px; color: #0f172a; white-space: normal !important; word-break: keep-all !important; margin-bottom: 6px; }
    .prod-item-desc { font-size: 13px; color: #475569; word-break: break-word !important; line-height: 1.5; }
    
    .res-card-box {
        position: relative;
        background: #f8fafc;
        border: 1px solid #e2e8f0;
        border-radius: 10px;
        padding: 12px 14px;
        text-align: center;
        transition: transform 0.2s ease, box-shadow 0.2s ease;
    }
    .res-card-box:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.05);
    }
    .res-card-box.inc { border-top: 4px solid #10b981; background: #f0fdf4; }
    .res-card-box.inc .res-card-val { color: #166534; }
    .res-card-box.exc { border-top: 4px solid #ef4444; background: #fef2f2; }
    .res-card-box.exc .res-card-val { color: #991b1b; }
    .res-card-box.pending { border-top: 4px solid #f59e0b; background: #fffbeb; }
    .res-card-box.pending .res-card-val { color: #92400e; }
    .res-card-box.dup { border-top: 4px solid #64748b; background: #f8fafc; }
    .res-card-box.dup .res-card-val { color: #334155; }
    
    .res-card-label { font-size: 11px; font-weight: 700; color: #64748b; text-transform: uppercase; margin-bottom: 4px; }
    .res-card-val { font-size: 22px; font-weight: 800; }

    div[data-testid="stColumn"] div.stButton > button {
        border: none !important;
        background: transparent !important;
        padding: 0 !important;
        width: 100% !important;
        box-shadow: none !important;
    }
</style>
""",
    unsafe_allow_html=True,
)

# --------------------------------------------------
# Session State 메모리 저장소 초기화
# --------------------------------------------------
if "tab1_result" not in st.session_state:
    st.session_state["tab1_result"] = None
if "tab2_result" not in st.session_state:
    st.session_state["tab2_result"] = None
if "tab3_result" not in st.session_state:
    st.session_state["tab3_result"] = None
if "tab_gie_result" not in st.session_state:
    st.session_state["tab_gie_result"] = None
if "tab_ct_result" not in st.session_state:
    st.session_state["tab_ct_result"] = None

if "uploader_key" not in st.session_state:
    st.session_state["uploader_key"] = 0
if "show_reset_msg" not in st.session_state:
    st.session_state["show_reset_msg"] = False
if "screened_history" not in st.session_state:
    st.session_state["screened_history"] = {}
if "radio_category" not in st.session_state:
    st.session_state["radio_category"] = None
if "active_dashboard_filter" not in st.session_state:
    st.session_state["active_dashboard_filter"] = "ALL"

def clear_screening_results():
    st.session_state["tab1_result"] = None
    st.session_state["tab2_result"] = None
    st.session_state["tab3_result"] = None
    st.session_state["tab_gie_result"] = None
    st.session_state["tab_ct_result"] = None
    st.session_state["active_dashboard_filter"] = "ALL"

def reset_to_home():
    clear_screening_results()
    st.session_state["radio_category"] = None

def clear_history():
    st.session_state["screened_history"] = {}
    st.session_state["uploader_key"] += 1  
    st.session_state["show_reset_msg"] = True  

# --------------------------------------------------
# 클릭 및 수동 수정 가능한 대시보드 함수
# --------------------------------------------------
def render_interactive_dashboard(df, key_prefix):
    total_cnt = len(df)
    inc_cnt = len(df[df["AI 판정"] == "Include (포함)"])
    exc_cnt = len(df[df["AI 판정"] == "Exclude (제외)"])
    pending_cnt = len(df[df["AI 판정"].str.contains("Full-text Screening Needed|Manual Review Needed", na=False)])
    dup_cnt = len(df[df["AI 판정"].str.contains("Duplicated", na=False)])

    c1, c2, c3, c4, c5 = st.columns(5)
    
    with c1:
        if st.button(f"전체_{total_cnt}", key=f"{key_prefix}_btn_all", use_container_width=True):
            st.session_state["active_dashboard_filter"] = "ALL"
        st.markdown(f'<div class="res-card-box"><div class="res-card-label">전체 대상</div><div class="res-card-val">{total_cnt}건</div></div>', unsafe_allow_html=True)
        
    with c2:
        if st.button(f"포함_{inc_cnt}", key=f"{key_prefix}_btn_inc", use_container_width=True):
            st.session_state["active_dashboard_filter"] = "INC"
        st.markdown(f'<div class="res-card-box inc"><div class="res-card-label">INCLUDE (포함)</div><div class="res-card-val">{inc_cnt}건</div></div>', unsafe_allow_html=True)

    with c3:
        if st.button(f"제외_{exc_cnt}", key=f"{key_prefix}_btn_exc", use_container_width=True):
            st.session_state["active_dashboard_filter"] = "EXC"
        st.markdown(f'<div class="res-card-box exc"><div class="res-card-label">EXCLUDE (제외)</div><div class="res-card-val">{exc_cnt}건</div></div>', unsafe_allow_html=True)

    with c4:
        if st.button(f"보류_{pending_cnt}", key=f"{key_prefix}_btn_pending", use_container_width=True):
            st.session_state["active_dashboard_filter"] = "PENDING"
        st.markdown(f'<div class="res-card-box pending"><div class="res-card-label">FULL-TEXT/MANUAL REVIEW NEEDED</div><div class="res-card-val">{pending_cnt}건</div></div>', unsafe_allow_html=True)

    with c5:
        if st.button(f"중복_{dup_cnt}", key=f"{key_prefix}_btn_dup", use_container_width=True):
            st.session_state["active_dashboard_filter"] = "DUP"
        st.markdown(f'<div class="res-card-box dup"><div class="res-card-label">DUPLICATED (중복)</div><div class="res-card-val">{dup_cnt}건</div></div>', unsafe_allow_html=True)

    current_filter = st.session_state.get("active_dashboard_filter", "ALL")
    filtered_df = df.copy()

    if current_filter == "INC":
        filtered_df = filtered_df[filtered_df["AI 판정"] == "Include (포함)"]
    elif current_filter == "EXC":
        filtered_df = filtered_df[filtered_df["AI 판정"] == "Exclude (제외)"]
        
        exclude_reasons = [
            "전체 제외 사유 보기",
            "Literature without human clinical data",
            "Irrelevant article",
            "Different indication",
            "Insufficient information",
            "Held by Taewoong Medical"
        ]
        selected_reason = st.selectbox(
            "세부 Exclude 사유 필터:", 
            options=exclude_reasons, 
            key=f"{key_prefix}_reason_filter"
        )
        if selected_reason != "전체 제외 사유 보기":
            filtered_df = filtered_df[filtered_df["Conclusion"].str.contains(selected_reason, na=False, case=False)]

    elif current_filter == "PENDING":
        filtered_df = filtered_df[filtered_df["AI 판정"].str.contains("Full-text Screening Needed|Manual Review Needed", na=False)]
    elif current_filter == "DUP":
        filtered_df = filtered_df[filtered_df["AI 판정"].str.contains("Duplicated", na=False)]

    st.markdown("<br>", unsafe_allow_html=True)

    edited_df = st.data_editor(
        filtered_df,
        key=f"{key_prefix}_editor_{current_filter}",
        hide_index=True,
        use_container_width=True,
        column_config={
            "AI 판정": st.column_config.SelectboxColumn(
                "AI 판정 (수동 수정 가능)",
                options=["Include (포함)", "Exclude (제외)", "Manual Review Needed", "Duplicated"],
                required=True
            ),
            "Conclusion": st.column_config.TextColumn("Conclusion (수동 수정 가능)", width="large"),
        }
    )

    excel_data = convert_df_to_excel(edited_df)
    st.download_button(
        "현재 목록 Excel(.xlsx) 다운로드",
        data=excel_data,
        file_name=f"screening_result_{current_filter.lower()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False
    )

# --------------------------------------------------
# 사이드바 UI 구성
# --------------------------------------------------
with st.sidebar:
    try:
        default_api_key = st.secrets["GEMINI_API_KEY"]
    except Exception:
        default_api_key = ""

    st.markdown(
        """
        <div style="font-size: 20px; font-weight: 700; color: #0F172A; margin-bottom: 2px;">
            Gemini API Key
        </div>
        <div style="font-size: 14px; font-weight: 400; color: #64748B; margin-bottom: 6px;">
            (미입력 시 클라우드 기본키 적용)
        </div>
        """,
        unsafe_allow_html=True
    )
    user_api_key = st.text_input(
        "Gemini API Key Input", type="password", label_visibility="collapsed"
    )
    api_key = user_api_key.strip() if user_api_key.strip() else default_api_key

    try:
        ncbi_api_key = st.secrets["NCBI_API_KEY"]
    except Exception:
        ncbi_api_key = ""

    if api_key:
        st.success("Gemini API Key 등록 완료")
    else:
        st.error("Gemini API Key가 필요합니다.")
        
    st.markdown("---")
    st.subheader("품목 선택")

    category_options = [
        "1. Biliary Stent",
        "2. Esophageal Stent",
        "3. Pyloric/Duodenal Stent",
        "4. Colonic Stent",
        "5. Drainage Stent",
    ]

    due_category = st.radio(
        "스크리닝할 카테고리를 선택하세요",
        options=category_options,
        index=(
            category_options.index(st.session_state["radio_category"])
            if st.session_state.get("radio_category") in category_options
            else None
        ),
        key="radio_category",
        on_change=clear_screening_results,
    )

    current_engine = st.session_state.get("engine_mode_seg", "PubMed Engine")
    sub_model = None

    if current_engine == "ClinicalTrials Engine":
        sub_model = None
    else:
        if due_category == "1. Biliary Stent":
            sub_model = st.selectbox(
                "세부 모델/유형을 선택하세요",
                options=[
                    "Niti-S Biliary Uncovered Stent",
                    "Niti-S Biliary Covered Stent",
                    "ComVi Biliary Stent",
                ],
                index=0,
                key="sb_biliary",
                on_change=clear_screening_results,
            )
        elif due_category == "2. Esophageal Stent":
            sub_model = st.selectbox(
                "세부 모델/유형을 선택하세요",
                options=["Niti-S Esophageal Covered Stent"],
                index=0,
                key="sb_esophageal",
                on_change=clear_screening_results,
            )
        elif due_category == "3. Pyloric/Duodenal Stent":
            sub_model = st.selectbox(
                "세부 모델/유형을 선택하세요",
                options=[
                    "Niti-S Pyloric/Duodenal Uncovered Stent",
                    "Niti-S Pyloric/Duodenal Covered Stent",
                    "ComVi Pyloric/Duodenal Stent",
                ],
                index=0,
                key="sb_pyloric",
                on_change=clear_screening_results,
            )
        elif due_category == "4. Colonic Stent":
            sub_model = st.selectbox(
                "세부 모델/유형을 선택하세요",
                options=[
                    "Niti-S Enteral Colonic Uncovered Stent",
                    "Niti-S Enteral Colonic Covered Stent",
                    "ComVi Enteral Colonic Stent",
                ],
                index=0,
                key="sb_colonic",
                on_change=clear_screening_results,
            )
        elif due_category == "5. Drainage Stent":
            sub_model = st.selectbox(
                "세부 모델/유형을 선택하세요",
                options=[
                    "Niti-S SPAXUS Stent",
                    "Niti-S Hot SPAXUS Stent",
                    "Niti-S Nagi Stent",
                ],
                index=0,
                key="sb_drainage",
                on_change=clear_screening_results,
            )

    st.markdown("---")
    history_cnt = len(st.session_state.get("screened_history", {}))
    st.caption(f"현재 누적 스크리닝 이력: **{history_cnt}건**")

    with st.expander("이전 스크리닝 결과 불러오기"):
        history_files = st.file_uploader(
            "과거 스크리닝 결과 파일 선택",
            type=["xlsx", "csv"],
            accept_multiple_files=True,
            key=f"history_csv_uploader_{st.session_state['uploader_key']}", 
        )
        if st.button("이력 메모리에 복원", use_container_width=True):
            if history_files:
                restored_count = 0
                for h_file in history_files:
                    try:
                        if h_file.name.endswith(".xlsx"):
                            h_df = pd.read_excel(h_file)
                        else:
                            try:
                                h_df = pd.read_csv(h_file, encoding="utf-8")
                            except UnicodeDecodeError:
                                h_df = pd.read_csv(h_file, encoding="cp949")

                        for _, row in h_df.iterrows():
                            identifier = None
                            if "PMID" in row and pd.notna(row["PMID"]):
                                identifier = str(row["PMID"]).replace(".0", "").strip()
                            elif "NCT 번호 (URL)" in row and pd.notna(row["NCT 번호 (URL)"]):
                                identifier = str(row["NCT 번호 (URL)"]).split('/')[-1].strip().lower()
                            elif "DOI / URL" in row and pd.notna(row["DOI / URL"]):
                                identifier = str(row["DOI / URL"]).strip().lower()
                            elif "논문 제목" in row and pd.notna(row["논문 제목"]):
                                identifier = str(row["논문 제목"]).strip().lower()
                            elif "임상시험 제목" in row and pd.notna(row["임상시험 제목"]):
                                identifier = str(row["임상시험 제목"]).strip().lower()

                            if identifier and identifier != "-":
                                cat = str(row.get("카테고리", "기존 이력"))
                                mod = str(row.get("세부 모델", "과거 파일"))
                                res = str(row.get("AI 판정", "Screened"))

                                if "screened_history" not in st.session_state:
                                    st.session_state["screened_history"] = {}

                                st.session_state["screened_history"][identifier] = {
                                    "category": cat,
                                    "sub_model": mod,
                                    "result": res,
                                }
                                restored_count += 1
                        st.success(f"총 {restored_count}건의 이력이 복원되었습니다!")
                    except Exception as e:
                        st.error(f"파일 읽기 오류 ({h_file.name}): {str(e)}")
            else:
                st.warning("복원할 파일을 선택하세요.")

    if st.button("이전 스크리닝 기록 초기화"):
        clear_history()

    if st.session_state.get("show_reset_msg", False):
        st.success("누적 기록 및 업로드 파일이 성공적으로 초기화되었습니다.")
        st.session_state["show_reset_msg"] = False  
        time.sleep(2)  
        st.rerun()    

    st.markdown("---")
    st.button(
        "HOME",
        type="secondary",
        use_container_width=True,
        on_click=reset_to_home,
    )

# --------------------------------------------------
# 🏠 1. 메인 홈 대시보드
# --------------------------------------------------
if not due_category:
    dept_info = "Development Department &nbsp;|&nbsp; Development 2nd Team"
    st.markdown(
        f"""<div class="hero-container">
<div class="hero-header-flex">
<div class="hero-tag">TAEWOONG MEDICAL CLINICAL EVALUATION PLATFORM</div>
<div class="dept-tag">{dept_info}</div>
</div>
<div class="hero-title">AI 문헌 스크리닝 시스템</div>
<div class="hero-subtitle">Medical Device Regulatory Compliance & Systematic Literature Review Powered by Gemini 3.6 Flash</div>
</div>""",
        unsafe_allow_html=True,
    )

    col_ov1, col_ov2, col_ov3 = st.columns([2.5, 1, 1.5])

    with col_ov1:
        with st.container(border=True):
            st.markdown(
                """
                <div class="card-title">TARGET PRODUCTS</div>
                <div class="card-value">Taewoong Medical’s Stent Product Lines</div>
                <div class="card-desc">아래 상자를 클릭하여 세부 라인업 및 제품 카탈로그 정보를 확인하세요.</div>
                """,
                unsafe_allow_html=True,
            )

            with st.expander("View Detailed Product Catalog"):
                c_tab1, c_tab2, c_tab3, c_tab4, c_tab5 = st.tabs([
                    "Biliary", "Esophageal", "Pyloric/Duodenal", "Colonic", "Drainage"
                ])

                with c_tab1:
                    st.markdown("#### **Niti-S & ComVi Biliary Stent**")
                    b_sub1, b_sub2, b_sub3 = st.tabs(["Uncovered Stent", "Covered Stent", "ComVi Stent"])
                    
                    with b_sub1:
                        biliary_uncovered_models = [
                            ("S-Type", "biliary_uncovered_s.png", "Niti-S Biliary Uncovered Stent [S-Type] is indicated for use in malignant strictures."),
                            ("D-Type", "biliary_uncovered_d.png", "Niti-S Biliary Uncovered Stent [D-Type] is indicated for use in malignant strictures."),
                            ("M-Type", "biliary_uncovered_m.png", "Niti-S Biliary Uncovered Stent [M-Type] is indicated for use in malignant strictures."),
                            ("LCD-Type", "biliary_uncovered_lcd.png", "Niti-S Biliary Uncovered Stent [LCD-Type] is indicated for use in malignant strictures."),
                        ]
                        with st.container(border=True):
                            for m_name, m_img, m_desc in biliary_uncovered_models:
                                c1, c2 = st.columns([1, 2.2])
                                with c1:
                                    if os.path.exists(m_img):
                                        st.image(m_img, use_container_width=True)
                                    else:
                                        st.caption(f"이미지 등록 필요: {m_img}")
                                with c2:
                                    st.markdown(
                                        f'<div class="prod-item-title">Niti-S Biliary Uncovered Stent [{m_name}]</div>'
                                        f'<div class="prod-item-desc">• {m_desc}</div>',
                                        unsafe_allow_html=True,
                                    )
                                st.divider()

                    with b_sub2:
                        biliary_covered_models = [
                            ("Full Covered-Type", "biliary_covered_full.png", "Niti-S Biliary Covered Stent [Full Covered-Type] is indicated for use in malignant and/or benign strictures."),
                            ("Both Bare-Type", "biliary_covered_bothbare.png", "Niti-S Biliary Covered Stent [Both Bare-Type] is indicated for use in malignant strictures."),
                            ("Giobor", "biliary_covered_giobor.png", "Niti-S Biliary Covered Stent [Giobor] is indicated for use in malignant strictures."),
                            ("Flare-Type", "biliary_covered_flare.png", "Niti-S Biliary Covered Stent [Flare-Type] is indicated for use in malignant and/or benign strictures."),
                            ("Kaffes", "biliary_covered_kaffes.png", "Niti-S Biliary Covered Stent [Kaffes] is indicated for use in malignant and/or benign strictures."),
                            ("Bumpy", "biliary_covered_bumpy.png", "Niti-S Biliary Covered Stent [Bumpy] is indicated for use in malignant and/or benign biliary strictures and benign pancreatic ductal strictures."),
                        ]
                        with st.container(border=True):
                            for m_name, m_img, m_desc in biliary_covered_models:
                                c1, c2 = st.columns([1, 2.2])
                                with c1:
                                    if os.path.exists(m_img):
                                        st.image(m_img, use_container_width=True)
                                    else:
                                        st.caption(f"이미지 등록 필요: {m_img}")
                                with c2:
                                    st.markdown(
                                        f'<div class="prod-item-title">Niti-S Biliary Covered Stent [{m_name}]</div>'
                                        f'<div class="prod-item-desc">• {m_desc}</div>',
                                        unsafe_allow_html=True,
                                    )
                                st.divider()

                    with b_sub3:
                        biliary_comvi_models = [
                            ("Full Covered-Type", "biliary_comvi_full.png", "ComVi Biliary Stent [Full Covered-Type] is indicated for use in malignant strictures."),
                            ("Both Bare-Type", "biliary_comvi_bothbare.png", "ComVi Biliary Stent [Both Bare-Type] is indicated for use in malignant strictures."),
                            ("End Bare-Type", "biliary_comvi_endbare.png", "ComVi Biliary Stent [End Bare-Type] is indicated for use in malignant strictures."),
                        ]
                        with st.container(border=True):
                            for m_name, m_img, m_desc in biliary_comvi_models:
                                c1, c2 = st.columns([1, 2.2])
                                with c1:
                                    if os.path.exists(m_img):
                                        st.image(m_img, use_container_width=True)
                                    else:
                                        st.caption(f"이미지 등록 필요: {m_img}")
                                with c2:
                                    st.markdown(
                                        f'<div class="prod-item-title">ComVi Biliary Stent [{m_name}]</div>'
                                        f'<div class="prod-item-desc">• {m_desc}</div>',
                                        unsafe_allow_html=True,
                                    )
                                st.divider()

                with c_tab2:
                    st.markdown("#### **Niti-S Esophageal Stent**")
                    e_sub1 = st.tabs(["Covered Stent"])[0]
                    with e_sub1:
                        esophageal_covered_models = [
                            ("Full Covered-Type", "esophageal_covered_full.png", "Niti-S Esophageal Covered Stent [Full Covered-Type] is indicated for use in malignant and/or refractory benign stricture and tracheoesophageal fistula."),
                            ("Cervical", "esophageal_covered_cervical.png", "Niti-S Esophageal Covered Stent [Cervical] is indicated for use in malignant strictures."),
                            ("Both Bare-Type", "esophageal_covered_bothbare.png", "Niti-S Esophageal Covered Stent [Both Bare-Type] is indicated for use in malignant strictures."),
                            ("Conio", "esophageal_covered_conio.png", "Niti-S Esophageal Covered Stent [Conio] is indicated for use in malignant and/or benign stricture and tracheoesophageal fistula."),
                            ("Anti Reflux-Type", "esophageal_covered_antireflux.png", "Niti-S Esophageal Covered Stent [Anti Reflux-Type] is indicated for use in malignant and/or benign stricture and tracheoesophageal fistula."),
                            ("Double Anti Reflux-Type", "esophageal_covered_doubleantireflux.png", "Niti-S Esophageal Covered Stent [Double Anti-Reflux-Type] is indicated for use in malignant strictures."),
                            ("Double-Type", "esophageal_covered_double.png", "Niti-S Esophageal Covered Stent [Double-Type] is indicated for use in malignant strictures."),
                            ("Beta-2", "esophageal_covered_beta2.png", "Niti-S Esophageal Covered Stent [Beta-2] is indicated for use in malignant and/or benign stricture and tracheoesophageal fistula."),
                        ]
                        with st.container(border=True):
                            for m_name, m_img, m_desc in esophageal_covered_models:
                                c1, c2 = st.columns([1, 2.2])
                                with c1:
                                    if os.path.exists(m_img):
                                        st.image(m_img, use_container_width=True)
                                    else:
                                        st.caption(f"이미지 등록 필요: {m_img}")
                                with c2:
                                    st.markdown(
                                        f'<div class="prod-item-title">Niti-S Esophageal Covered Stent [{m_name}]</div>'
                                        f'<div class="prod-item-desc">• {m_desc}</div>',
                                        unsafe_allow_html=True,
                                    )
                                st.divider()

                with c_tab3:
                    st.markdown("#### **Niti-S & ComVi Pyloric/Duodenal Stent**")
                    p_sub1, p_sub2, p_sub3 = st.tabs(["Uncovered Stent", "Covered Stent", "ComVi Stent"])
                    
                    with p_sub1:
                        pyloric_uncovered_models = [
                            ("D-Type", "pyloric_uncovered_d.png", "Niti-S Pyloric/Duodenal Uncovered Stent [D-Type] is indicated for use in intrinsic and/or extrinsic malignant stricture."),
                        ]
                        with st.container(border=True):
                            for m_name, m_img, m_desc in pyloric_uncovered_models:
                                c1, c2 = st.columns([1, 2.2])
                                with c1:
                                    if os.path.exists(m_img):
                                        st.image(m_img, use_container_width=True)
                                    else:
                                        st.caption(f"이미지 등록 필요: {m_img}")
                                with c2:
                                    st.markdown(
                                        f'<div class="prod-item-title">Niti-S Pyloric/Duodenal Uncovered Stent [{m_name}]</div>'
                                        f'<div class="prod-item-desc">• {m_desc}</div>',
                                        unsafe_allow_html=True,
                                    )
                                st.divider()

                    with p_sub2:
                        pyloric_covered_models = [
                            ("Full Covered-Type", "pyloric_covered_full.png", "Niti-S Pyloric/Duodenal Covered Stent [Full Covered-Type] is indicated for use in intrinsic and/or extrinsic malignant and/or benign stricture."),
                            ("Both Bare-Type", "pyloric_covered_bothbare.png", "Niti-S Pyloric/Duodenal Covered Stent [Both Bare-Type] is indicated for use in intrinsic and/or extrinsic malignant stricture."),
                            ("End Bare-Type", "pyloric_covered_endbare.png", "Niti-S Pyloric/Duodenal Covered Stent [End Bare-Type] is indicated for use in intrinsic and/or extrinsic malignant stricture."),
                        ]
                        with st.container(border=True):
                            for m_name, m_img, m_desc in pyloric_covered_models:
                                c1, c2 = st.columns([1, 2.2])
                                with c1:
                                    if os.path.exists(m_img):
                                        st.image(m_img, use_container_width=True)
                                    else:
                                        st.caption(f"이미지 등록 필요: {m_img}")
                                with c2:
                                    st.markdown(
                                        f'<div class="prod-item-title">Niti-S Pyloric/Duodenal Covered Stent [{m_name}]</div>'
                                        f'<div class="prod-item-desc">• {m_desc}</div>',
                                        unsafe_allow_html=True,
                                    )
                                st.divider()

                    with p_sub3:
                        pyloric_comvi_models = [
                            ("Flare-Type", "pyloric_comvi_flare.png", "ComVi Pyloric/Duodenal Stent [Flare-Type] is indicated for use in intrinsic and/or extrinsic malignant stricture."),
                            ("Both Bare-Type", "pyloric_comvi_bothbare.png", "ComVi Pyloric/Duodenal Stent [Both Bare-Type] is indicated for use in intrinsic and/or extrinsic malignant stricture."),
                        ]
                        with st.container(border=True):
                            for m_name, m_img, m_desc in pyloric_comvi_models:
                                c1, c2 = st.columns([1, 2.2])
                                with c1:
                                    if os.path.exists(m_img):
                                        st.image(m_img, use_container_width=True)
                                    else:
                                        st.caption(f"이미지 등록 필요: {m_img}")
                                with c2:
                                    st.markdown(
                                        f'<div class="prod-item-title">ComVi Pyloric/Duodenal Stent [{m_name}]</div>'
                                        f'<div class="prod-item-desc">• {m_desc}</div>',
                                        unsafe_allow_html=True,
                                    )
                                st.divider()

                with c_tab4:
                    st.markdown("#### **Niti-S & ComVi Enteral Colonic Stent**")
                    col_sub1, col_sub2, col_sub3 = st.tabs(["Uncovered Stent", "Covered Stent", "ComVi Stent"])
                    
                    with col_sub1:
                        colonic_uncovered_models = [
                            ("S-Type", "colonic_uncovered_s.png", "Niti-S Enteral Colonic Uncovered Stent [S-Type] is indicated for use in colon stricture caused by intrinsic and/or extrinsic malignant stricture."),
                            ("D-Type", "colonic_uncovered_d.png", "Niti-S Enteral Colonic Uncovered Stent [D-Type] is indicated for use in colon stricture caused by intrinsic and/or extrinsic malignant stricture."),
                        ]
                        with st.container(border=True):
                            for m_name, m_img, m_desc in colonic_uncovered_models:
                                c1, c2 = st.columns([1, 2.2])
                                with c1:
                                    if os.path.exists(m_img):
                                        st.image(m_img, use_container_width=True)
                                    else:
                                        st.caption(f"이미지 등록 필요: {m_img}")
                                with c2:
                                    st.markdown(
                                        f'<div class="prod-item-title">Niti-S Enteral Colonic Uncovered Stent [{m_name}]</div>'
                                        f'<div class="prod-item-desc">• {m_desc}</div>',
                                        unsafe_allow_html=True,
                                    )
                                st.divider()

                    with col_sub2:
                        colonic_covered_models = [
                            ("Full Covered-Type", "colonic_covered_full.png", "Niti-S Enteral Colonic Covered Stent [Full Covered-Type] is indicated for use in colon stricture caused by intrinsic and/or extrinsic malignant and/or benign stricture."),
                            ("Both Bare-Type", "colonic_covered_bothbare.png", "Niti-S Enteral Colonic Covered Stent [Both Bare-Type] is indicated for use in colon stricture caused by intrinsic and/or extrinsic malignant stricture."),
                            ("End Bare-Type", "colonic_covered_endbare.png", "Niti-S Enteral Colonic Covered Stent [End Bare-Type] is indicated for use in colon stricture caused by intrinsic and/or extrinsic malignant stricture."),
                        ]
                        with st.container(border=True):
                            for m_name, m_img, m_desc in colonic_covered_models:
                                c1, c2 = st.columns([1, 2.2])
                                with c1:
                                    if os.path.exists(m_img):
                                        st.image(m_img, use_container_width=True)
                                    else:
                                        st.caption(f"이미지 등록 필요: {m_img}")
                                with c2:
                                    st.markdown(
                                        f'<div class="prod-item-title">Niti-S Enteral Colonic Covered Stent [{m_name}]</div>'
                                        f'<div class="prod-item-desc">• {m_desc}</div>',
                                        unsafe_allow_html=True,
                                    )
                                st.divider()

                    with col_sub3:
                        colonic_comvi_models = [
                            ("Both Bare-Type", "colonic_comvi_bothbare.png", "ComVi Enteral Colonic Stent [Both Bare-Type] is indicated for use in colon stricture caused by intrinsic and/or extrinsic malignant stricture."),
                        ]
                        with st.container(border=True):
                            for m_name, m_img, m_desc in colonic_comvi_models:
                                c1, c2 = st.columns([1, 2.2])
                                with c1:
                                    if os.path.exists(m_img):
                                        st.image(m_img, use_container_width=True)
                                    else:
                                        st.caption(f"이미지 등록 필요: {m_img}")
                                with c2:
                                    st.markdown(
                                        f'<div class="prod-item-title">ComVi Enteral Colonic Stent [{m_name}]</div>'
                                        f'<div class="prod-item-desc">• {m_desc}</div>',
                                        unsafe_allow_html=True,
                                    )
                                st.divider()

                with c_tab5:
                    st.markdown("#### **Niti-S Drainage Stent**")
                    d_sub1, d_sub2, d_sub3 = st.tabs(["SPAXUS Stent", "Hot SPAXUS Stent", "Nagi Stent"])
                    
                    with d_sub1:
                        spaxus_models = [
                            ("SPAXUS", "drainage_spaxus.png", "Niti-S SPAXUS™ Stent is indicated for transgastric or transduodenal drainage of a pancreatic pseudocyst or a walled off necrosis or a gallbladder or the biliary tract."),
                        ]
                        with st.container(border=True):
                            for m_name, m_img, m_desc in spaxus_models:
                                c1, c2 = st.columns([1, 2.2])
                                with c1:
                                    if os.path.exists(m_img):
                                        st.image(m_img, use_container_width=True)
                                    else:
                                        st.caption(f"이미지 등록 필요: {m_img}")
                                with c2:
                                    st.markdown(
                                        f'<div class="prod-item-title">Niti-S SPAXUS Stent</div>'
                                        f'<div class="prod-item-desc">• {m_desc}</div>',
                                        unsafe_allow_html=True,
                                    )
                                st.divider()

                    with d_sub2:
                        hot_spaxus_models = [
                            ("Hot SPAXUS", "drainage_hot_spaxus.png", "Niti-S Hot SPAXUS™ Stent is indicated for transgastric or transduodenal drainage of a pancreatic pseudocyst or a walled off necrosis or a gallbladder or the biliary tract."),
                        ]
                        with st.container(border=True):
                            for m_name, m_img, m_desc in hot_spaxus_models:
                                c1, c2 = st.columns([1, 2.2])
                                with c1:
                                    if os.path.exists(m_img):
                                        st.image(m_img, use_container_width=True)
                                    else:
                                        st.caption(f"이미지 등록 필요: {m_img}")
                                with c2:
                                    st.markdown(
                                        f'<div class="prod-item-title">Niti-S Hot SPAXUS Stent</div>'
                                        f'<div class="prod-item-desc">• {m_desc}</div>',
                                        unsafe_allow_html=True,
                                    )
                                st.divider()

                    with d_sub3:
                        nagi_models = [
                            ("Nagi", "drainage_nagi.png", "Niti-S Nagi™ Stent is indicated for drainage of a pancreatic pseudocyst through a transgastric or transduodenal approach."),
                        ]
                        with st.container(border=True):
                            for m_name, m_img, m_desc in nagi_models:
                                c1, c2 = st.columns([1, 2.2])
                                with c1:
                                    if os.path.exists(m_img):
                                        st.image(m_img, use_container_width=True)
                                    else:
                                        st.caption(f"이미지 등록 필요: {m_img}")
                                with c2:
                                    st.markdown(
                                        f'<div class="prod-item-title">Niti-S Nagi Stent</div>'
                                        f'<div class="prod-item-desc">• {m_desc}</div>',
                                        unsafe_allow_html=True,
                                    )
                                st.divider()

    with col_ov2:
        with st.container(border=True):
            st.markdown(
                """
                <div class="card-title">AI PIPELINE</div>
                <div class="card-value">Gemini 3.6 Flash</div>
                <div class="card-desc">AI 기반 문헌 스크리닝</div>
                <br>
                """,
                unsafe_allow_html=True,
            )
    with col_ov3:
        with st.container(border=True):
            st.markdown(
                """
                <div class="card-title">REGULATORY GOAL</div>
                <div class="card-value">Standardization and Automation of Literature Review</div>
                <div class="card-desc">일관성 및 추적성을 확보한 스크리닝 기록</div>
                <br>
                """,
                unsafe_allow_html=True,
            )
    st.stop()

# --------------------------------------------------
# 세부 모델별 프롬프트 및 PICO 키워드 세팅
# --------------------------------------------------
if due_category == "1. Biliary Stent":
    include_criteria = """1. Text availability: Full text (Original articles, Reviews, Case reports/series 모두 포함)
2. Species: Human (not animal, artificial simulation)
3. Patient population: Adult patients, irrespective of gender
4. Clinical Conditions: Malignant biliary obstruction/stricture, Benign biliary obstruction/stricture, Benign pancreatic duct stricture
5. Intervention: Biliary SEMS (Uncovered or Covered). Specific Taewoong Medical models: Niti-S (S, D, M, LCD, Full Covered, Both Bare, Giobor, Flare, Kaffes, Bumpy), ComVi (Full Covered, Both Bare, End Bare)
6. Comparators: Surgery, Plastic stent, Balloon dilation, or competitor SEMS (e.g., WallFlex, Evolution, EGIS, Bonastent, Hanarostent)
7. Outcomes: Stent patency, Decreased bilirubin, Technical/Clinical success, Complications, Stent removal (for benign cases)"""
    exclude_criteria = """1. Species: Not human beings (animal test, artificial simulation, in vitro test)
2. Different indication: Non-biliary/pancreatic target areas only (e.g., vascular, esophageal, colonic, tracheal)
3. Irrelevant articles: Articles not related to biliary/pancreatic luminal stenting or stricture management (e.g., EUS-CDS, EUS-HGS primary LAMS procedures, RFA combined therapies, vascular reconstructions)
4. Non-study publications: Editorials, letters, comments, study protocols (단, Review 및 Case report는 제외하지 않음)
5. Insufficient Information: Valid information relevant to performance and/or safety is limited.
6. Held by Taewoong: This article is already held by Taewoong Medical."""

    if sub_model == "Niti-S Biliary Uncovered Stent":
        default_p, default_i, default_c, default_o = "Biliary obstruction\nBiliary stricture\nMalignant biliary stricture\nMalignant biliary obstruction", "Self-expandable metallic stent\nSelf-expandable metal stent\nSEMS\nTaewoong\nNiti-S\nUncovered stent", "Surgery\nPlastic stent\nBalloon dilation\nSelf-expandable metallic stent\nSelf-expandable metal stent\nSEMS\nUncovered stent\nEvolution\nWallFlex\nEGIS\nBonastent\nHanarostent", "Stent patency\nDecreased bilirubin"
        add_search_queries = [
            'Taewoong AND Niti-S AND Biliary AND "S type"',
            'Taewoong AND Niti-S AND Biliary AND "D type"',
            'Taewoong AND Niti-S AND Biliary AND "M type"',
            'Taewoong AND Niti-S AND Biliary AND "LCD type"'
        ]
    elif sub_model == "Niti-S Biliary Covered Stent":
        default_p, default_i, default_c, default_o = "Biliary obstruction\nBiliary stricture\nBenign biliary stricture\nBenign biliary obstruction\nMalignant biliary stricture\nMalignant biliary obstruction\nBenign pancreatic duct stricture", "Self-expandable metallic stent\nSelf-expandable metal stent\nSEMS\nTaewoong\nNiti-S\nCovered stent", "Surgery\nPlastic stent\nBalloon dilation\nSelf-expandable metallic stent\nSelf-expandable metal stent\nSEMS\nCovered stent\nEvolution\nWallFlex\nEGIS\nBonastent\nHanarostent", "Stent patency\nDecreased bilirubin\nRemoval"
        add_search_queries = [
            'Taewoong AND Niti-S AND Biliary AND "Full covered type"',
            'Taewoong AND Niti-S AND Biliary AND "Both bare type"',
            'Taewoong AND Niti-S AND Biliary AND Giobor',
            'Taewoong AND Niti-S AND Biliary AND "Flare type"',
            'Taewoong AND Niti-S AND Biliary AND Kaffes',
            'Taewoong AND Niti-S AND Biliary AND Bumpy'
        ]
    elif sub_model == "ComVi Biliary Stent":
        default_p, default_i, default_c, default_o = "Biliary obstruction\nBiliary stricture\nMalignant biliary stricture\nMalignant biliary obstruction", "Self-expandable metallic stent\nSelf-expandable metal stent\nSEMS\nTaewoong\nComVi\nCovered stent", "Surgery\nPlastic stent\nBalloon dilation\nSelf-expandable metallic stent\nSelf-expandable metal stent\nSEMS\nCovered stent\nEvolution\nWallFlex\nEGIS\nBonastent\nHanarostent", "Stent patency\nDecreased bilirubin"
        add_search_queries = [
            'Taewoong AND ComVi AND Biliary AND "Full covered type"',
            'Taewoong AND ComVi AND Biliary AND "Both bare type"',
            'Taewoong AND ComVi AND Biliary AND "End bare type"'
        ]
    else:
        default_p, default_i, default_c, default_o = "Biliary obstruction\nBiliary stricture\nMalignant biliary stricture\nMalignant biliary obstruction\nBenign biliary obstruction\nBenign biliary stricture\nBenign pancreatic duct stricture", "Self-expandable metallic stent\nSelf-expandable metal stent\nSEMS\nTaewoong\nNiti-S\nComVi\nUncovered stent\nCovered stent", "Surgery\nPlastic stent\nBalloon dilation\nSelf-expandable metallic stent\nSelf-expandable metal stent\nSEMS\nCovered stent\nUncovered stent\nEvolution\nWallFlex\nEGIS\nBonastent\nHanarostent", "Stent patency\nDecreased bilirubin\nRemoval"
        add_search_queries = ['Taewoong AND Biliary AND Stent']

elif due_category == "2. Esophageal Stent":
    include_criteria = """1. Text availability: Full text (Original articles, Reviews, Case reports/series 모두 포함)
2. Species: Human (not animal, artificial simulation)
3. Patient population: Adult patients, irrespective of gender
4. Clinical Conditions: Esophageal stricture/obstruction (Malignant or Benign), Refractory benign esophageal stricture, Tracheoesophageal fistula (TEF / TE fistula)
5. Intervention: Esophageal SEMS, Covered type. Specific Taewoong Medical models: Niti-S Esophageal (Full covered, Cervical, Both bare type, Conio, Anti reflux, Double anti reflux, Double type, Beta-2)
6. Comparators: Surgery, Plastic stent, Balloon dilation, or competitor SEMS (WallFlex, Ultraflex, Evolution, Hanarostent, Aixstent, EGIS, Bonastent, Micro-Tech)
7. Outcomes: Stent patency, Dysphagia improvement, Fistula closure, Removal (in benign strictures)"""
    exclude_criteria = """1. Species: Not human beings (animal test, artificial simulation, in vitro test)
2. Different indication: Non-esophageal target areas only (e.g., pure systemic/chemotherapy outcomes, non-stricture indications)
3. Irrelevant articles: Articles not related to esophageal stenting, stricture dilation, or TE fistula management
4. Non-study publications: Editorials, letters, comments, study protocols (단, Review 및 Case report는 제외하지 않음)
5. Insufficient Information: Valid information relevant to performance and/or safety is limited.
6. Held by Taewoong: This article is already held by Taewoong Medical."""
    default_p, default_i, default_c, default_o = "Esophageal stricture\nEsophageal obstruction\nMalignant esophageal stricture\nMalignant esophageal obstruction\nBenign esophageal stricture\nRefractory benign esophageal stricture\nBenign esophgeal obstruction\nTracheoesophageal fistula", "Self-expandable metallic stent\nSelf-expandable metal stent\nSEMS\nTaewoong\nNiti-S\nCovered stent", "Surgery\nPlastic stent\nBalloon dilation\nSelf-expandable metallic stent\nSelf-expandable metal stent\nSEMS\nCovered stent\nWallFlex\nUltraflex\nEvolution\nHanarostent\nAixstent\nEGIS\nBonastent\nMicro-tech", "Stent patency\nDysphagia improvement\nFistula closure\nRemoval"
    add_search_queries = [
        'Taewoong AND Niti-S AND Esophageal AND "Full covered type"',
        'Taewoong AND Niti-S AND Esophageal AND Cervical',
        'Taewoong AND Niti-S AND Esophageal AND "Both bare type"',
        'Taewoong AND Niti-S AND Esophageal AND Conio',
        'Taewoong AND Niti-S AND Esophageal AND "Anti reflux"',
        'Taewoong AND Niti-S AND Esophageal AND "Double anti reflux"',
        'Taewoong AND Niti-S AND Esophageal AND "Double type"',
        'Taewoong AND Niti-S AND Esophageal AND Beta-2'
    ]

elif due_category == "3. Pyloric/Duodenal Stent":
    include_criteria = """1. Text availability: Full text (Original articles, Reviews, Case reports/series 모두 포함)
2. Species: Human (not animal, artificial simulation)
3. Patient population: Adult patients, irrespective of gender
4. Clinical Conditions: Pyloric/Duodenal stricture or obstruction, Gastric Outlet Obstruction (GOO), Malignant or Benign
5. Intervention: Pyloric/Duodenal SEMS, Uncovered or Covered type. Specific Taewoong Medical models: Niti-S Pyloric/Duodenal (D-Type, Full Covered, Both Bare, End Bare), ComVi Pyloric/Duodenal (Flare-Type, Both Bare)
6. Comparators: Surgery, Plastic stent, Balloon dilation, or competitor SEMS (WallFlex, WallFlex Soft, Hanarostent, Evolution, EGIS, Bonastent)
7. Outcomes: Stent patency, Obstruction relief/resolution/improvement, GOOSS score / Oral intake, Technical/Clinical success, Complications, Stent removal"""
    exclude_criteria = """1. Species: Not human beings (animal test, artificial simulation, in vitro test)
2. Different indication: Non-pyloric/duodenal target areas only
3. Irrelevant articles: Articles not related to pyloric/duodenal stenting or GOO management
4. Non-study publications: Editorials, letters, comments, study protocols
5. Insufficient Information: Valid information relevant to performance and/or safety is limited.
6. Held by Taewoong: This article is already held by Taewoong Medical."""
    if sub_model == "Niti-S Pyloric/Duodenal Uncovered Stent":
        default_p, default_i, default_c, default_o = "Pyloric stricture\nPyloric obstruction\nDuodenal stricture\nDuodenal obstruction\nGastric outlet obstruction\nMalignant pyloric stricture\nMalignant pyloric obstruction\nMalignant duodenal stricture\nMalignant duodenal obstruction", "Self-expandable metallic stent\nSelf-expandable metal stent\nSEMS\nTaewoong\nNiti-S\nUncovered stent", "Surgery\nPlastic stent\nBalloon dilation\nSelf-expandable metallic stent\nSelf-expandable metal stent\nSEMS\nUncovered stent\nWallFlex\nWallFlex Soft\nHanarostent\nEvolution\nEGIS\nBonastent", "Stent patency\nObstruction relief\nObstruction resolution\nObstruction improvement"
        add_search_queries = [
            'Taewoong AND Niti-S AND (Pyloric OR Duodenal) AND "D type"'
        ]
    elif sub_model == "Niti-S Pyloric/Duodenal Covered Stent":
        default_p, default_i, default_c, default_o = "Pyloric stricture\nPyloric obstruction\nDuodenal stricture\nDuodenal obstruction\nGastric outlet obstruction\nMalignant pyloric stricture\nMalignant pyloric obstruction\nMalignant duodenal stricture\nMalignant duodenal obstruction\nBenign pyloric stricture\nBenign pyloric obstruction\nBenign duodenal stricture\nBenign duodenal obstruction", "Self-expandable metallic stent\nSelf-expandable metal stent\nSEMS\nTaewoong\nNiti-S\nCovered stent", "Surgery\nPlastic stent\nBalloon dilation\nSelf-expandable metallic stent\nSelf-expandable metal stent\nSEMS\nCovered stent\nWallFlex\nWallFlex Soft\nHanarostent\nEvolution\nEGIS\nBonastent", "Stent patency\nObstruction relief\nObstruction resolution\nObstruction improvement\nRemoval"
        add_search_queries = [
            'Taewoong AND Niti-S AND (Pyloric OR Duodenal) AND "Full covered type"',
            'Taewoong AND Niti-S AND (Pyloric OR Duodenal) AND "Both bare type"',
            'Taewoong AND Niti-S AND (Pyloric OR Duodenal) AND "End bare type"'
        ]
    elif sub_model == "ComVi Pyloric/Duodenal Stent":
        default_p, default_i, default_c, default_o = "Pyloric stricture\nPyloric obstruction\nDuodenal stricture\nDuodenal obstruction\nGastric outlet obstruction\nMalignant pyloric stricture\nMalignant pyloric obstruction\nMalignant duodenal stricture\nMalignant duodenal obstruction", "Self-expandable metallic stent\nSelf-expandable metal stent\nSEMS\nTaewoong\nComVi\nCovered stent", "Surgery\nPlastic stent\nBalloon dilation\nSelf-expandable metallic stent\nSelf-expandable metal stent\nSEMS\nCovered stent\nWallFlex\nWallFlex Soft\nHanarostent\nEvolution\nEGIS\nBonastent", "Stent patency\nObstruction relief\nObstruction resolution\nObstruction improvement"
        add_search_queries = [
            'Taewoong AND ComVi AND (Pyloric OR Duodenal) AND "Flare type"',
            'Taewoong AND ComVi AND (Pyloric OR Duodenal) AND "Both bare type"'
        ]
    else:
        default_p, default_i, default_c, default_o = "Pyloric stricture\nDuodenal stricture\nGastric outlet obstruction\nMalignant pyloric stricture\nBenign pyloric stricture", "SEMS\nTaewoong\nNiti-S\nComVi\nCovered stent\nUncovered stent", "Surgery\nPlastic stent\nBalloon dilation\nSEMS\nWallFlex", "Stent patency\nObstruction relief\nRemoval"
        add_search_queries = ['Taewoong AND (Pyloric OR Duodenal) AND Stent']

elif due_category == "4. Colonic Stent":
    include_criteria = """1. Text availability: Full text (Original articles, Reviews, Case reports/series 모두 포함)
2. Species: Human (not animal, artificial simulation)
3. Patient population: Adult patients, irrespective of gender
4. Clinical Conditions: Colonic/Colorectal stricture or obstruction (Malignant or Benign)
5. Intervention: Colonic SEMS, Uncovered or Covered type. Specific Taewoong Medical models: Niti-S Enteral Colonic (S-Type, D-Type, Full Covered, Both Bare, End Bare), ComVi Enteral Colonic (Both Bare-Type)
6. Comparators: Surgery, Plastic stent, Balloon dilation, or competitor SEMS (WallFlex, WallFlex Soft, Hanarostent, Micro-Tech, Bonastent)
7. Outcomes: Stent patency, Obstruction relief/resolution/improvement, Technical/Clinical success, Complications, Stent removal"""
    exclude_criteria = """1. Species: Not human beings (animal test, artificial simulation, in vitro test)
2. Different indication: Non-colonic target areas only
3. Irrelevant articles: Articles not related to colonic stenting or colorectal obstruction management
4. Non-study publications: Editorials, letters, comments, study protocols
5. Insufficient Information: Valid information relevant to performance and/or safety is limited.
6. Held by Taewoong: This article is already held by Taewoong Medical."""
    if sub_model == "Niti-S Enteral Colonic Uncovered Stent":
        default_p, default_i, default_c, default_o = "Colonic stricture\nColonic obstruction\nColorectal stricture\nColorectal obstruction\nMalignant colonic stricture\nMalignant colonic obstruction\nMalignant colorectal stricture\nMalignant colorectal obstruction", "Self-expandable metallic stent\nSelf-expandable metal stent\nSEMS\nTaewoong\nNiti-S\nUncovered stent", "Surgery\nPlastic stent\nBalloon dilation\nSelf-expandable metallic stent\nSelf-expandable metal stent\nSEMS\nUncovered stent\nWallFlex\nWallFlex Soft\nHanarostent\nMicro-Tech\nBonastent", "Stent patency\nObstruction relief\nObstruction resolution\nObstruction improvement"
        add_search_queries = [
            'Taewoong AND Niti-S AND Colonic AND "S type"',
            'Taewoong AND Niti-S AND Colonic AND "D type"'
        ]
    elif sub_model == "Niti-S Enteral Colonic Covered Stent":
        default_p, default_i, default_c, default_o = "Colonic stricture\nColonic obstruction\nColorectal stricture\nColorectal obstruction\nMalignant colonic stricture\nMalignant colonic obstruction\nMalignant colorectal stricture\nMalignant colorectal obstruction\nBenign colonic stricture\nBenign colonic obstruction\nBenign colorectal stricture\nBenign colorectal obstruction", "Self-expandable metallic stent\nSelf-expandable metal stent\nSEMS\nTaewoong\nNiti-S\nCovered stent", "Surgery\nPlastic stent\nBalloon dilation\nSelf-expandable metallic stent\nSelf-expandable metal stent\nSEMS\nCovered stent\nWallFlex\nWallFlex Soft\nHanarostent\nMicro-Tech\nBonastent", "Stent patency\nObstruction relief\nObstruction resolution\nObstruction improvement\nRemoval"
        add_search_queries = [
            'Taewoong AND Niti-S AND Colonic AND "Full covered type"',
            'Taewoong AND Niti-S AND Colonic AND "Both bare type"',
            'Taewoong AND Niti-S AND Colonic AND "End bare type"'
        ]
    elif sub_model == "ComVi Enteral Colonic Stent":
        default_p, default_i, default_c, default_o = "Colonic stricture\nColonic obstruction\nColorectal stricture\nColorectal obstruction\nMalignant colonic stricture\nMalignant colonic obstruction\nMalignant colorectal stricture\nMalignant colorectal obstruction", "Self-expandable metallic stent\nSelf-expandable metal stent\nSEMS\nTaewoong\nComVi\nCovered stent", "Surgery\nPlastic stent\nBalloon dilation\nSelf-expandable metallic stent\nSelf-expandable metal stent\nSEMS\nCovered stent\nWallFlex\nWallFlex Soft\nHanarostent\nMicro-Tech\nBonastent", "Stent patency\nObstruction relief\nObstruction resolution\nObstruction improvement"
        add_search_queries = [
            'Taewoong AND ComVi AND Colonic AND "Both bare type"'
        ]
    else:
        default_p, default_i, default_c, default_o = "Colonic stricture\nColorectal obstruction\nMalignant colonic stricture\nBenign colonic stricture", "SEMS\nTaewoong\nNiti-S\nComVi\nUncovered stent\nCovered stent", "Surgery\nPlastic stent\nBalloon dilation\nSEMS\nWallFlex", "Stent patency\nObstruction relief\nRemoval"
        add_search_queries = ['Taewoong AND Colonic AND Stent']

elif due_category == "5. Drainage Stent":
    include_criteria = """1. Text availability: Full text (Original articles, Reviews, Case reports/series 모두 포함)
2. Species: Human (not animal, artificial simulation)
3. Patient population: Adult patients, irrespective of gender
4. Clinical Conditions: Pancreatic pseudocyst, Walled-off necrosis (WON) / Pancreatic necrosis, Gallbladder drainage (Cholecystitis) / Biliary tract drainage, Transgastric or transduodenal drainage indications
5. Intervention: Lumen-apposing metal stents (LAMS) or EUS-guided drainage stents. Specific Taewoong Medical models: Niti-S Nagi, Niti-S SPAXUS, Niti-S Hot SPAXUS
6. Comparators: Surgery, Percutaneous drainage, Plastic double-pigtail stents, or competitor LAMS (e.g., AXIOS / Hot AXIOS)
7. Outcomes: Technical/Clinical success rate, Drainage efficacy, Resolution of pseudocyst/necrosis, Complications, Removal rate"""
    exclude_criteria = """1. Species: Not human beings (animal test, artificial simulation, in vitro test)
2. Different indication: Non-drainage target indications
3. Irrelevant articles: Articles not related to transluminal/EUS-guided drainage or LAMS management
4. Non-study publications: Editorials, letters, comments, study protocols
5. Insufficient Information: Valid information relevant to performance and/or safety is limited.
6. Held by Taewoong: This article is already held by Taewoong Medical."""
    if sub_model == "Niti-S SPAXUS Stent":
        default_p, default_i, default_c, default_o = "Pancreatic pseudocyst\nWalled off necrosis\nGallbladder\nBiliary tract", "Self-expandable metallic stent\nSEMS\nLumen apposing metal stent\nLAMS\nEUS gallbladder drainage\nEUS choledochoduodenostomy\nTaewoong\nNiti-S\nSPAXUS", "Surgery\nPercutaneous drainage\nSEMS\nLAMS\nAXIOS\nHot AXIOS", "Pancreatic pseudocyst drainage\nWalled off necrosis drainage\nGallbladder drainage\nBiliary tract drainage"
        add_search_queries = ['Taewoong AND Niti-S AND SPAXUS']
    elif sub_model == "Niti-S Hot SPAXUS Stent":
        default_p, default_i, default_c, default_o = "Pancreatic pseudocyst\nWalled off necrosis\nGallbladder\nBiliary tract", "SEMS\nLumen apposing metal stent\nLAMS\nEUS gallbladder\nEUS choledochoduodenostomy\nElectrocautery delivery system\nHot delivery\nTaewoong\nNiti-S\nHot SPAXUS", "Surgery\nPercutaneous drainage\nSEMS\nLAMS\nAXIOS\nHot AXIOS", "Pancreatic pseudocyst drainage\nWalled off necrosis drainage\nGallbladder drainage\nBiliary tract drainage"
        add_search_queries = ['Taewoong AND Niti-S AND "Hot SPAXUS"']
    elif sub_model == "Niti-S Nagi Stent":
        default_p, default_i, default_c, default_o = "Pancreatic pseudocyst", "SEMS\nLumen apposing metal stent\nLAMS\nBiflanged metal stent\nBFMS\nTaewoong\nNiti-S\nNagi", "Surgery\nPercutaneous drainage\nSEMS\nLAMS\nBiflanged metal stent\nBFMS\nAXIOS\nHot AXIOS", "Pancreatic pseudocyst drainage"
        add_search_queries = ['Taewoong AND Niti-S AND Nagi']
    else:
        default_p, default_i, default_c, default_o = "Pancreatic pseudocyst\nWalled off necrosis\nGallbladder", "SEMS\nLAMS\nTaewoong\nNiti-S\nComVi", "Surgery\nPercutaneous drainage\nLAMS\nAXIOS", "Drainage"
        add_search_queries = ['Taewoong AND Niti-S AND Drainage']

else:
    include_criteria = "Include All Relevant Clinical Papers"
    exclude_criteria = "Exclude Non-Clinical/Irrelevant Papers"
    default_p, default_i, default_c, default_o = "Obstructive Jaundice\nBiliary Stricture", "Biliary Stent\nSEMS", "Surgery\nPlastic stent", "Technical success\nClinical success"
    add_search_queries = ['Taewoong AND Stent']

# --------------------------------------------------
# ✨ 2단계 세그먼티드 컨트롤 메뉴
# --------------------------------------------------
sub_model_html = (
    f'<div style="font-size: 14px; font-weight: 500; color: #94a3b8; margin-top: 4px;">{sub_model}</div>'
    if sub_model else ""
)

st.markdown(
    f"""
<div class="selected-category-box">
    <div class="selected-category-label">Selected Category</div>
    <div class="selected-category-title">{due_category}</div>
    {sub_model_html}
</div>
""",
    unsafe_allow_html=True,
)

target_engine = st.segmented_control(
    "", options=["PubMed Engine", "GIE Journal Engine", "ClinicalTrials Engine"], default="PubMed Engine", key="engine_mode_seg",
)
st.markdown("<br>", unsafe_allow_html=True)

if target_engine == "PubMed Engine":
    selected_mode = st.segmented_control(
        "", options=["PubMed PICO 자동 검색", "PMID 리스트 업로드", "단일 PMID 입력"], default="PubMed PICO 자동 검색", key="pubmed_sub_mode_seg",
    )
elif target_engine == "GIE Journal Engine":
    selected_mode = st.segmented_control(
        "", options=["GIE RIS 파일 일괄 스크리닝"], default="GIE RIS 파일 일괄 스크리닝", key="gie_sub_mode_seg",
    )
else:
    selected_mode = st.segmented_control(
        "", options=["ClinicalTrials 자동 검색"], default="ClinicalTrials 자동 검색", key="ct_sub_mode_seg",
    )

st.markdown("<br>", unsafe_allow_html=True)

# --------------------------------------------------
# MODE 1: 단일 PMID 테스트
# --------------------------------------------------
if selected_mode == "단일 PMID 입력":
    single_pmid = st.text_input("PubMed PMID 번호를 입력하세요 (예: 31234567)")
    if st.button("단일 PMID 스크리닝 실행"):
        if not api_key:
            st.error("API Key가 설정되지 않았습니다!")
        elif not single_pmid:
            st.error("PMID 번호를 입력해 주세요!")
        else:
            with st.spinner("PubMed 공식 API에서 논문 정보 조회 중..."):
                title, abs_text, pmcid, status = fetch_pubmed_by_pmid(single_pmid, ncbi_api_key)

            pmid_url = f"https://pubmed.ncbi.nlm.nih.gov/{single_pmid.strip()}/"

            article_content = ""
            eval_source = ""

            if pmcid:
                st.info(f"Open Access 논문 확인됨 (PMCID: {pmcid}). 전문(Full-text)을 가져옵니다.")
                full_text = fetch_pmc_fulltext(pmcid, ncbi_api_key)
                if full_text:
                    eval_source = "Full-text (Open Access)"
                    article_content = f"[Title]\n{title}\n\n[Full-text Body]\n{full_text}"
                else:
                    eval_source = "Abstract Only"
                    article_content = f"[Title]\n{title}\n\n[Abstract]\n{abs_text}"
            elif abs_text:
                eval_source = "Abstract Only"
                article_content = f"[Title]\n{title}\n\n[Abstract]\n{abs_text}"
            else:
                eval_source = "No Data"

            if eval_source == "No Data":
                st.warning(f"**논문 제목:** {title if title else '제목 없음'}")
                st.info(f"원문 직접 링크: [{pmid_url}]({pmid_url})")
                st.error(f"판정: **Manual Review Needed (Abstract Missing)**")
                st.caption(f"Reason: **Insufficient information:** Abstract/Full-text is unavailable. Manual full-text review is required.")
            else:
                st.success(f"**논문 제목:** {title}")
                st.caption(f"평가 기준: **{eval_source}**")
                
                genai.configure(api_key=api_key)
                model = genai.GenerativeModel("gemini-3.6-flash")
                prompt = generate_prompt(due_category, include_criteria, exclude_criteria, title, article_content)

                ans_text, err = call_gemini_with_retry(model, prompt)
                if ans_text:
                    st.success("AI 스크리닝 판정 완료!")
                    st.markdown(ans_text)
                else:
                    st.error(f"AI 통신 에러 발생: {err}")

# --------------------------------------------------
# MODE 2: PMID 리스트 파일 업로드
# --------------------------------------------------
elif selected_mode == "PMID 리스트 업로드":
    uploaded_file = st.file_uploader("PMID가 적힌 파일 업로드 ('PMID' 열 필수)", type=["xlsx", "csv"])
    if st.button("PMID 일괄 스크리닝 실행"):
        if not api_key: st.error("API Key가 설정되지 않았습니다!")
        elif not uploaded_file: st.error("파일을 업로드해 주세요!")
        else:
            try:
                if uploaded_file.name.endswith(".xlsx"):
                    df = pd.read_excel(uploaded_file)
                else:
                    try: df = pd.read_csv(uploaded_file, encoding="utf-8")
                    except Exception: df = pd.read_csv(uploaded_file, encoding="cp949")
            except Exception as e:
                st.error(f"파일 읽기 오류: {str(e)}")
                df = pd.DataFrame()

            if "PMID" not in df.columns:
                st.error("업로드 파일 안에 'PMID' 열이 존재해야 합니다.")
            else:
                df = df.loc[:, ~df.columns.str.contains("^Unnamed")]
                if "No" in df.columns: df = df.drop(columns=["No"])

                genai.configure(api_key=api_key)
                model = genai.GenerativeModel("gemini-3.6-flash")

                titles, abstracts, results, conclusions, pubmed_urls, eval_sources = [], [], [], [], [], []
                progress_bar = st.progress(0)
                status_text = st.empty()
                total = len(df)

                for idx, row in df.iterrows():
                    pmid = str(row["PMID"]).replace(".0", "").strip()
                    title, abs_text, pmcid, status = fetch_pubmed_by_pmid(pmid, ncbi_api_key)
                    pmid_url = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"
                    pubmed_urls.append(pmid_url)

                    content_for_ai = ""
                    eval_source = ""

                    if pmcid:
                        status_text.markdown(f"**[{idx+1}/{total}] PMC 원문 수집 & 정밀 분석 중...** (PMID: {pmid})")
                        full_text = fetch_pmc_fulltext(pmcid, ncbi_api_key)
                        if full_text:
                            eval_source = "Full-text (Open Access)"
                            content_for_ai = f"[Title]\n{title}\n\n[Full-text Body]\n{full_text}" 
                        else:
                            eval_source = "Abstract Only"
                            content_for_ai = f"[Title]\n{title}\n\n[Abstract]\n{abs_text}"
                    elif abs_text:
                        status_text.markdown(f"**[{idx+1}/{total}] 초록(Abstract) 분석 중...** (PMID: {pmid})")
                        eval_source = "Abstract Only"
                        content_for_ai = f"[Title]\n{title}\n\n[Abstract]\n{abs_text}"
                    else:
                        status_text.markdown(f"**[{idx+1}/{total}] 데이터 확보 불가, 예외 처리 중...** (PMID: {pmid})")
                        eval_source = "No Data"

                    if eval_source == "No Data":
                        titles.append(title if title else "조회 실패")
                        abstracts.append("No Abstract or Full-text Available")
                        results.append("Manual Review Needed")
                        conclusions.append(to_unicode_bold(f"Insufficient information: No abstract or open-access full-text available. Manual retrieval and review required."))
                        eval_sources.append(eval_source)
                    else:
                        identifier = pmid
                        if identifier in st.session_state.get("screened_history", {}):
                            prev_info = st.session_state["screened_history"][identifier]
                            prev_mod = prev_info["sub_model"]
                            titles.append(title)
                            abstracts.append(abs_text[:150] + "..." if abs_text else "원문 확인 (요약 생략)")
                            results.append(f"Duplicated (이전중복: {prev_mod})")
                            conclusions.append(f"Duplicate literature previously screened in [{prev_mod}] step.")
                            eval_sources.append(eval_source)
                        else:
                            titles.append(title)
                            abstracts.append(abs_text[:150] + "..." if abs_text else "원문 확인 (요약 생략)")

                            prompt = generate_prompt(due_category, include_criteria, exclude_criteria, title, content_for_ai)
                            ans, err = call_gemini_with_retry(model, prompt)

                            if ans:
                                res_label = "Include (포함)" if "Include" in ans and "Exclude" not in ans.split("판정:")[1] else "Exclude (제외)"
                                results.append(res_label)
                                if "Conclusion:" in ans:
                                    raw_conclusion = ans.split("Conclusion:")[-1].strip()
                                else:
                                    raw_conclusion = ans.split("\n\n")[-1].replace("Conclusion:", "").strip()
                                conclusions.append(to_unicode_bold(raw_conclusion))
                                
                                if "screened_history" not in st.session_state:
                                    st.session_state["screened_history"] = {}
                                st.session_state["screened_history"][identifier] = {
                                    "category": due_category, "sub_model": sub_model, "result": res_label
                                }
                            else:
                                results.append("Error")
                                conclusions.append(f"AI 에러: {err}")
                            eval_sources.append(eval_source)

                    progress_bar.progress((idx + 1) / total)
                    time.sleep(0.12 if ncbi_api_key else 0.35)

                status_text.empty(); progress_bar.empty()

                df["카테고리"] = due_category
                df["세부 모델"] = sub_model
                df["논문 제목"] = titles
                df["평가 기준"] = eval_sources
                df["초록 요약"] = abstracts
                df["AI 판정"] = results
                df["Conclusion"] = conclusions
                df["PubMed Link"] = pubmed_urls

                df.insert(0, "No", range(1, len(df) + 1))
                df.index = df.index + 1
                st.session_state["tab2_result"] = df

    if st.session_state["tab2_result"] is not None:
        st.success(f"[{due_category} - {sub_model}] 일괄 스크리닝 결과")
        render_interactive_dashboard(st.session_state["tab2_result"], key_prefix="tab2")

# --------------------------------------------------
# MODE 3: PubMed PICO 키워드 자동 검색 & 스크리닝
# --------------------------------------------------
elif selected_mode == "PubMed PICO 자동 검색":
    st.subheader(f"PubMed PICO 키워드")
    st.caption("선택하신 품목 및 세부 모델에 맞춰 P, I, C, O 키워드가 자동으로 세팅되었습니다.")

    col_pico1, col_pico2 = st.columns(2)
    with col_pico1:
        p_val = st.text_area("P (Patient / Population / Problem)", value=default_p, height=140)
        i_val = st.text_area("I (Intervention)", value=default_i, height=140)
    with col_pico2:
        c_val = st.text_area("C (Comparison)", value=default_c, height=140)
        o_val = st.text_area("O (Outcome)", value=default_o, height=140)

    st.markdown("---")
    st.subheader("PICO 검색 조합 선택")
    st.caption("원하시는 퀵 세팅 버튼을 누르거나, 아래 체크박스를 통해 검색에 포함할 항목을 직접 조정하세요.")

    if "pico_preset" not in st.session_state:
        st.session_state["pico_preset"] = "P + I"

    col_q1, col_q2, col_q3, col_q4, col_q5 = st.columns(5)
    with col_q1:
        if st.button("P + I", use_container_width=True):
            st.session_state["pico_preset"] = "P + I"
    with col_q2:
        if st.button("P + O", use_container_width=True):
            st.session_state["pico_preset"] = "P + O"
    with col_q3:
        if st.button("P + C + O", use_container_width=True):
            st.session_state["pico_preset"] = "P + C + O"
    with col_q4:
        if st.button("직접 조합 (Custom)", use_container_width=True):
            st.session_state["pico_preset"] = "Custom"

    current_preset = st.session_state["pico_preset"]

    default_chk_p = current_preset in ["P + I", "P + O", "P + C + O"]
    default_chk_i = current_preset in ["P + I"]
    default_chk_c = current_preset in ["P + C + O"]
    default_chk_o = current_preset in ["P + O", "P + C + O"]
    default_chk_add = current_preset == "I 단독"

    st.markdown("<br>", unsafe_allow_html=True)

    col_ck1, col_ck2, col_ck3, col_ck4, col_ck5 = st.columns(5)
    with col_ck1:
        use_p = st.checkbox("P (Patient / Population / Problem)", value=default_chk_p)
    with col_ck2:
        use_i = st.checkbox("I (Intervention)", value=default_chk_i)
    with col_ck3:
        use_c = st.checkbox("C (Comparison)", value=default_chk_c)
    with col_ck4:
        use_o = st.checkbox("O (Outcome)", value=default_chk_o)
    with col_ck5:
        use_add_i = st.checkbox("Additional Search I", value=default_chk_add)

    def sync_query_input():
        st.session_state["custom_direct_query_input"] = st.session_state["sb_add_query_preset"]

    selected_target_direct_query = None
    if use_add_i:
        st.markdown("<br>", unsafe_allow_html=True)
        
        if "custom_direct_query_input" not in st.session_state:
            st.session_state["custom_direct_query_input"] = add_search_queries[0]

        selected_q_preset = st.selectbox(
            "개별 Additional Search 정밀 쿼리 프리셋 선택:",
            options=add_search_queries,
            index=0,
            key="sb_add_query_preset",
            on_change=sync_query_input
        )
        
        selected_target_direct_query = st.text_input(
            "실행될 쿼리문 (필요 시 직접 수정 가능):",
            key="custom_direct_query_input"
        )

    st.markdown("---")
    st.subheader("문헌 검색 기간(연/월) 및 추출 개수 설정")

    col_s1, col_s2, col_e1, col_e2 = st.columns(4)
    with col_s1: start_year = st.number_input("시작 연도", min_value=1990, max_value=2026, value=2026)
    with col_s2: start_month = st.number_input("시작 월", min_value=1, max_value=12, value=1)
    with col_e1: end_year = st.number_input("종료 연도", min_value=1990, max_value=2026, value=2026)
    with col_e2: end_month = st.number_input("종료 월", min_value=1, max_value=12, value=8)

    col_opt1, col_opt2 = st.columns([1, 1])
    with col_opt1: fetch_all_toggle = st.checkbox("검색된 전체 논문 수집 (개수 제한 없음)", value=False)
    with col_opt2:
        if not fetch_all_toggle: max_limit = st.number_input("가져올 최대 논문 수", min_value=1, max_value=2000, value=20)
        else: max_limit = 0; st.info("조건에 맞는 PubMed의 전체 PMID를 가져옵니다.")

    if st.button("PICO 다중 조합 검색 및 AI 스크리닝 실행"):
        found_pmids = []
        used_query = ""

        if use_add_i and selected_target_direct_query:
            date_range_label = f"{start_year}년 {start_month:02d}월 ~ {end_year}년 {end_month:02d}월"
            with st.spinner(f"PubMed에서 [{selected_target_direct_query}] 단독 쿼리 실행 중..."):
                found_pmids, used_query = search_pubmed_pmids_pico(
                    start_year=start_year, start_month=start_month, end_year=end_year, end_month=end_month,
                    fetch_all=fetch_all_toggle, max_results=max_limit, ncbi_api_key=ncbi_api_key, direct_query=selected_target_direct_query
                )

        else:
            if not use_p and not use_i and not use_c and not use_o:
                use_p, use_i = True, True
                st.warning("선택된 요소가 없어 기본값 (P + I) 조합으로 자동 지정되었습니다.")

            target_p = p_val if use_p else ""
            target_i = i_val if use_i else ""
            target_c = c_val if use_c else ""
            target_o = o_val if use_o else ""

            if not (target_p.strip() or target_i.strip() or target_c.strip() or target_o.strip()):
                st.error("최소한 하나 이상의 유효한 키워드가 선택되어 있어야 합니다!")
            else:
                date_range_label = f"{start_year}년 {start_month:02d}월 ~ {end_year}년 {end_month:02d}월"
                with st.spinner(f"PubMed에서 [{date_range_label}] 기간의 PICO 조합 조건으로 검색 중..."):
                    found_pmids, used_query = search_pubmed_pmids_pico(
                        p_text=target_p, i_text=target_i, c_text=target_c, o_text=target_o,
                        start_year=start_year, start_month=start_month, end_year=end_year, end_month=end_month,
                        fetch_all=fetch_all_toggle, max_results=max_limit, ncbi_api_key=ncbi_api_key
                    )

        if not found_pmids and used_query:
            st.warning(f"[{date_range_label}] 기간 및 선택하신 조건에 부합하는 PubMed 논문이 없습니다.")
            st.info(f"생성된 조합 쿼리:\n`{used_query}`")
        elif found_pmids:
            st.success(f"**[{date_range_label}]** 검색 결과, 총 **{len(found_pmids)}건**의 PMID가 추출되었습니다!")
            with st.expander("자동 생성된 PubMed 조합 쿼리식 확인", expanded=True): st.code(used_query, language="sql")

            pmid_summary = ", ".join(found_pmids[:10])
            st.write(f"**추출된 PMID 목록 (상위 10개):** {pmid_summary}" + (" ... (이하 생략)" if len(found_pmids) > 10 else ""))
            st.markdown("---")

            genai.configure(api_key=api_key)
            model = genai.GenerativeModel("gemini-3.6-flash")

            auto_df = pd.DataFrame({"PMID": found_pmids})
            titles, abstracts, results, conclusions, pubmed_urls, eval_sources = [], [], [], [], [], []
            progress_bar = st.progress(0)
            status_text = st.empty()
            total = len(auto_df)

            for idx, row in auto_df.iterrows():
                pmid = str(row["PMID"])
                title, abs_text, pmcid, status = fetch_pubmed_by_pmid(pmid, ncbi_api_key)
                pmid_url = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"
                pubmed_urls.append(pmid_url)

                content_for_ai = ""
                eval_source = ""

                if pmcid:
                    status_text.markdown(f"**[{idx+1}/{total}] PMC 원문 수집 & 정밀 분석 중...** (PMID: {pmid})")
                    full_text = fetch_pmc_fulltext(pmcid, ncbi_api_key)
                    if full_text:
                        eval_source = "Full-text (Open Access)"
                        content_for_ai = f"[Title]\n{title}\n\n[Full-text Body]\n{full_text}" 
                    else:
                        eval_source = "Abstract Only"
                        content_for_ai = f"[Title]\n{title}\n\n[Abstract]\n{abs_text}"
                elif abs_text:
                    status_text.markdown(f"**[{idx+1}/{total}] 초록(Abstract) 분석 중...** (PMID: {pmid})")
                    eval_source = "Abstract Only"
                    content_for_ai = f"[Title]\n{title}\n\n[Abstract]\n{abs_text}"
                else:
                    status_text.markdown(f"**[{idx+1}/{total}] 데이터 확보 불가, 예외 처리 중...** (PMID: {pmid})")
                    eval_source = "No Data"

                if eval_source == "No Data":
                    titles.append(title if title else "조회 실패")
                    abstracts.append("No Abstract or Full-text Available")
                    results.append("Manual Review Needed")
                    conclusions.append(to_unicode_bold(f"Insufficient information: No abstract or open-access full-text available. Manual retrieval and review required."))
                    eval_sources.append(eval_source)
                else:
                    identifier = pmid
                    if identifier in st.session_state.get("screened_history", {}):
                        prev_info = st.session_state["screened_history"][identifier]
                        prev_mod = prev_info["sub_model"]
                        titles.append(title)
                        abstracts.append(abs_text[:150] + "..." if abs_text else "원문 확인 (요약 생략)")
                        results.append(f"Duplicated (이전중복: {prev_mod})")
                        conclusions.append(f"Duplicate literature previously screened in [{prev_mod}] step.")
                        eval_sources.append(eval_source)
                    else:
                        titles.append(title)
                        abstracts.append(abs_text[:150] + "..." if abs_text else "원문 확인 (요약 생략)")

                        prompt = generate_prompt(due_category, include_criteria, exclude_criteria, title, content_for_ai)
                        ans, err = call_gemini_with_retry(model, prompt)

                        if ans:
                            res_label = "Include (포함)" if "Include" in ans and "Exclude" not in ans.split("판정:")[1] else "Exclude (제외)"
                            results.append(res_label)
                            if "Conclusion:" in ans:
                                raw_conclusion = ans.split("Conclusion:")[-1].strip()
                            else:
                                raw_conclusion = ans.split("\n\n")[-1].replace("Conclusion:", "").strip()
                            conclusions.append(to_unicode_bold(raw_conclusion))
                            
                            if "screened_history" not in st.session_state:
                                st.session_state["screened_history"] = {}
                            st.session_state["screened_history"][identifier] = {
                                "category": due_category, "sub_model": sub_model, "result": res_label
                            }
                        else:
                            results.append("Error")
                            conclusions.append(f"AI 에러: {err}")
                        eval_sources.append(eval_source)

            progress_bar.progress((idx + 1) / total)
            time.sleep(0.12 if ncbi_api_key else 0.35)

        status_text.empty(); progress_bar.empty()

        auto_df["카테고리"] = due_category
        auto_df["세부 모델"] = sub_model
        auto_df["논문 제목"] = titles
        auto_df["평가 기준"] = eval_sources
        auto_df["초록 요약"] = abstracts
        auto_df["AI 판정"] = results
        auto_df["Conclusion"] = conclusions
        auto_df["PubMed Link"] = pubmed_urls
        auto_df.insert(0, "No", range(1, len(auto_df) + 1))
        st.session_state["tab3_result"] = auto_df

    if st.session_state["tab3_result"] is not None:
        st.success(f"[{due_category} - {sub_model}] PICO 기반 자동 스크리닝 완료 결과")
        render_interactive_dashboard(st.session_state["tab3_result"], key_prefix="tab3")

# --------------------------------------------------
# MODE 4: GIE RIS 파일 전용 일괄 AI 스크리닝
# --------------------------------------------------
elif selected_mode == "GIE RIS 파일 일괄 스크리닝":
    st.subheader("GIE Journal RIS 파일 업로드 스크리닝")
    st.caption("giejournal.org 접속 -> Advanced Search 실행 -> [Export Citation] -> RIS 파일 다운로드 -> 아래 업로드 창에 드래그 & 드롭")

    gie_file = st.file_uploader("GIE에서 Export한 RIS 파일(.ris)을 업로드하세요", type=["ris", "txt"])

    if st.button("GIE RIS 파일 AI 스크리닝 실행"):
        if not api_key: st.error("API Key가 설정되지 않았습니다!")
        elif not gie_file: st.error("GIE RIS 파일을 업로드해 주세요!")
        else:
            try: content = gie_file.getvalue().decode("utf-8")
            except UnicodeDecodeError: content = gie_file.getvalue().decode("cp949", errors="ignore")

            try: entries = rispy.loads(content)
            except Exception as e: st.error(f"RIS 파일 구조 파싱 실패: {str(e)}"); entries = []

            if not entries: st.error("업로드하신 RIS 파일에서 추출된 논문 정보가 없습니다.")
            else:
                genai.configure(api_key=api_key)
                model = genai.GenerativeModel("gemini-3.6-flash")

                titles, abstracts, results, conclusions, doi_list, eval_sources = [], [], [], [], [], []
                progress_bar = st.progress(0)
                status_text = st.empty()
                total = len(entries)

                for idx, entry in enumerate(entries):
                    title = entry.get("title", entry.get("primary_title", "제목 없음"))
                    abstract_text = entry.get("abstract", "").strip()
                    doi = entry.get("doi", entry.get("url", "-")).strip()

                    status_text.markdown(f"**[{idx+1}/{total}] GIE 초록 AI 분석 진행 중...** ({title[:30]}...)")
                    identifier = doi.lower() if doi and doi != "-" else title.lower()

                    doi_list.append(doi)
                    eval_source = "Abstract Only" if abstract_text else "No Data"

                    if eval_source == "No Data":
                        titles.append(title)
                        abstracts.append("No Abstract Available")
                        results.append("Manual Review Needed")
                        conclusions.append(to_unicode_bold(f"Insufficient information: Abstract text is missing in the GIE RIS file. Manual full-text review is required."))
                        eval_sources.append(eval_source)
                    else:
                        if identifier in st.session_state.get("screened_history", {}):
                            prev_info = st.session_state["screened_history"][identifier]
                            prev_mod = prev_info["sub_model"]
                            titles.append(title)
                            abstracts.append(abstract_text[:150] + "...")
                            results.append(f"Duplicated (이전중복: {prev_mod})")
                            conclusions.append(f"Duplicate literature previously screened in [{prev_mod}] step.")
                            eval_sources.append(eval_source)
                        else:
                            titles.append(title)
                            abstracts.append(abstract_text[:150] + "...")

                            prompt = generate_prompt(due_category, include_criteria, exclude_criteria, title, abstract_text)
                            ans, err = call_gemini_with_retry(model, prompt)

                            if ans:
                                res_label = "Include (포함)" if "Include" in ans and "Exclude" not in ans.split("판정:")[1] else "Exclude (제외)"
                                results.append(res_label)
                                if "Conclusion:" in ans:
                                    raw_conclusion = ans.split("Conclusion:")[-1].strip()
                                else:
                                    raw_conclusion = ans.split("\n\n")[-1].replace("Conclusion:", "").strip()
                                conclusions.append(to_unicode_bold(raw_conclusion))
                                
                                if "screened_history" not in st.session_state:
                                    st.session_state["screened_history"] = {}
                                st.session_state["screened_history"][identifier] = {
                                    "category": due_category, "sub_model": sub_model, "result": res_label
                                }
                            else:
                                results.append("Error")
                                conclusions.append(f"AI 에러: {err}")
                            eval_sources.append(eval_source)

                    progress_bar.progress((idx + 1) / total)
                    time.sleep(0.3)

                status_text.empty(); progress_bar.empty()

                res_df = pd.DataFrame({
                    "No": range(1, len(entries) + 1),
                    "카테고리": due_category,
                    "세부 모델": sub_model,
                    "DOI / URL": doi_list,
                    "논문 제목": titles,
                    "평가 기준": eval_sources,
                    "초록 요약": abstracts,
                    "AI 판정": results,
                    "Conclusion": conclusions
                })
                st.session_state["tab_gie_result"] = res_df

    if st.session_state["tab_gie_result"] is not None and selected_mode == "GIE RIS 파일 일괄 스크리닝":
        st.success(f"[{due_category} - {sub_model}] GIE RIS 스크리닝 완료 결과")
        render_interactive_dashboard(st.session_state["tab_gie_result"], key_prefix="gie")

# --------------------------------------------------
# MODE 5: ClinicalTrials.gov 전용 API 자동 스크리닝
# --------------------------------------------------
elif selected_mode == "ClinicalTrials 자동 검색":
    st.subheader("Clinical Trials Registry (NCT) 검색")
    st.caption("세부 모델 구분 없이 선택하신 품목 카테고리 전체의 임상시험 데이터를 통합 검색합니다.")

    if due_category == "1. Biliary Stent":
        ct_default_cond = '("Biliary obstruction" OR "Biliary stricture")'
        ct_default_intr = '("Self-expandable metal stent" OR "SEMS" OR "Biliary stent")'
    elif due_category == "2. Esophageal Stent":
        ct_default_cond = '("Esophageal obstruction" OR "Esophageal stricture" OR "Tracheoesophageal fistula")'
        ct_default_intr = '("Self-expandable metal stent" OR "SEMS" OR "Esophageal stent")'
    elif due_category == "3. Pyloric/Duodenal Stent":
        ct_default_cond = '("Pyloric obstruction" OR "Duodenal obstruction" OR "Pyloric stricture" OR "Duodenal stricture")'
        ct_default_intr = '("Self-expandable metal stent" OR "SEMS" OR "Pyloric stent" OR "Duodenal stent")'
    elif due_category == "4. Colonic Stent":
        ct_default_cond = '("Colonic obstruction" OR "Colonic stricture")'
        ct_default_intr = '("Self-expandable metal stent" OR "SEMS" OR "Colonic stent")'
    elif due_category == "5. Drainage Stent":
        ct_default_cond = '("Pancreatic pseudocyst" OR "Walled off necrosis" OR "Gallbladder" OR "Biliary tract")'
        ct_default_intr = '("Lumen apposing metal stent" OR "LAMS" OR "Drainage stent")'
    else:
        ct_default_cond = '"Biliary stricture"'
        ct_default_intr = '"Stent"'

    col_ct1, col_ct2 = st.columns(2)
    with col_ct1:
        cond_val = st.text_input("Condition/disease", value=ct_default_cond)
    with col_ct2:
        intr_val = st.text_input("Intervention/treatment", value=ct_default_intr)

    col_f1, col_f2 = st.columns(2)
    with col_f1:
        status_options = {
            "Recruiting (모집 중)": "RECRUITING",
            "Completed (완료됨)": "COMPLETED",
            "Active, not recruiting (활성, 모집안함)": "ACTIVE_NOT_RECRUITING",
            "Not yet recruiting (모집 예정)": "NOT_YET_RECRUITING",
            "Terminated (조기 종료)": "TERMINATED"
        }
        selected_statuses = st.multiselect(
            "임상 진행 상태 (Status)", 
            options=list(status_options.keys()), 
            default=[],
            placeholder="미선택 시 전체 상태(All) 검색"
        )
        api_status_filters = [status_options[k] for k in selected_statuses]
        
    with col_f2:
        type_options = {
            "Interventional (중재적 연구)": "INTERVENTIONAL",
            "Observational (관찰 연구)": "OBSERVATIONAL"
        }
        selected_types = st.multiselect(
            "연구 유형 (Study Type)", 
            options=list(type_options.keys()),
            default=[],
            placeholder="미선택 시 전체 유형(All) 검색"
        )
        api_type_filters = [type_options[k] for k in selected_types]

    col_opt1, col_opt2 = st.columns([1, 1])
    with col_opt1:
        fetch_all_ct_toggle = st.checkbox("검색된 전체 임상시험 수집 (개수 제한 없음)", value=False)
    with col_opt2:
        if not fetch_all_ct_toggle:
            max_limit = st.number_input("가져올 최대 임상시험 수", min_value=1, max_value=2000, value=20)
        else:
            max_limit = 0
            st.info("조건에 부합하는 전체 임상시험을 페이지네이션으로 수집합니다.")

    if st.button("ClinicalTrials 검색 및 AI 스크리닝 실행"):
        if not api_key: st.error("Gemini API Key가 설정되지 않았습니다!")
        elif not cond_val.strip() and not intr_val.strip(): st.error("질환명이나 중재시술 중 하나는 반드시 입력해야 합니다!")
        else:
            with st.spinner("ClinicalTrials.gov 공식 API에서 조건에 맞는 데이터를 수집 중..."):
                studies, used_query = search_clinicaltrials(
                    cond_val, intr_val, 
                    status_filters=api_status_filters, 
                    type_filters=api_type_filters, 
                    fetch_all=fetch_all_ct_toggle,
                    max_results=max_limit
                )

            if not studies:
                st.warning("검색 조건에 일치하는 임상시험을 찾을 수 없습니다.")
                st.info(f"생성된 API 요청 URL:\n{used_query}")
            else:
                st.success(f"총 **{len(studies)}건**의 진행/완료된 임상시험 데이터가 추출되었습니다!")
                with st.expander("자동 생성된 API 쿼리 URL 확인", expanded=True):
                    st.code(used_query, language="http")

                genai.configure(api_key=api_key)
                model = genai.GenerativeModel("gemini-3.6-flash")

                titles, abstracts, results, conclusions, urls, eval_sources, statuses = [], [], [], [], [], [], []
                progress_bar = st.progress(0)
                status_text = st.empty()
                total = len(studies)

                for idx, (nct_id, title, summary, status) in enumerate(studies):
                    status_text.markdown(f"**[{idx+1}/{total}] 임상시험 브리핑(Summary) AI 분석 중...** ({nct_id})")
                    
                    nct_url = f"https://clinicaltrials.gov/study/{nct_id}"
                    identifier = nct_id.lower()
                    
                    urls.append(nct_url)
                    statuses.append(status)
                    
                    eval_source = "Abstract Only" if summary.strip() else "No Data"

                    if eval_source == "No Data":
                        titles.append(title)
                        abstracts.append("No Summary Available")
                        results.append("Manual Review Needed")
                        conclusions.append(to_unicode_bold(f"Insufficient information: No brief summary available for {nct_id}."))
                        eval_sources.append(eval_source)
                    else:
                        if identifier in st.session_state.get("screened_history", {}):
                            prev_info = st.session_state["screened_history"][identifier]
                            prev_mod = prev_info["sub_model"]
                            titles.append(title)
                            abstracts.append(summary[:150] + "...")
                            results.append(f"Duplicated (이전중복: {prev_mod})")
                            conclusions.append(f"Duplicate literature previously screened in [{prev_mod}] step.")
                            eval_sources.append(eval_source)
                        else:
                            titles.append(title)
                            abstracts.append(summary[:150] + "...")

                            article_content = f"[NCT ID]: {nct_id}\n[Status]: {status}\n[Title]\n{title}\n\n[Brief Summary]\n{summary}"
                            prompt = generate_prompt(due_category, include_criteria, exclude_criteria, title, article_content)
                            
                            ans, err = call_gemini_with_retry(model, prompt)

                            if ans:
                                res_label = "Include (포함)" if "Include" in ans and "Exclude" not in ans.split("판정:")[1] else "Exclude (제외)"
                                results.append(res_label)
                                if "Conclusion:" in ans:
                                    raw_conclusion = ans.split("Conclusion:")[-1].strip()
                                else:
                                    raw_conclusion = ans.split("\n\n")[-1].replace("Conclusion:", "").strip()
                                conclusions.append(to_unicode_bold(raw_conclusion))
                                
                                if "screened_history" not in st.session_state:
                                    st.session_state["screened_history"] = {}
                                st.session_state["screened_history"][identifier] = {
                                    "category": due_category, "sub_model": sub_model, "result": res_label
                                }
                            else:
                                results.append("Error")
                                conclusions.append(f"AI 에러: {err}")
                            eval_sources.append(eval_source)

                    progress_bar.progress((idx + 1) / total)
                    time.sleep(0.35) 

                status_text.empty()
                progress_bar.empty()

                res_df = pd.DataFrame({
                    "No": range(1, len(studies) + 1),
                    "카테고리": due_category,
                    "세부 모델": sub_model,
                    "NCT 번호 (URL)": urls,
                    "임상 진행 상태": statuses,
                    "임상시험 제목": titles,
                    "평가 기준": eval_sources,
                    "요약 (Summary)": abstracts,
                    "AI 판정": results,
                    "Conclusion": conclusions
                })
                st.session_state["tab_ct_result"] = res_df

    if st.session_state.get("tab_ct_result") is not None and selected_mode == "ClinicalTrials 자동 검색":
        st.success(f"[{due_category} - {sub_model}] ClinicalTrials.gov 스크리닝 완료 결과")
        render_interactive_dashboard(st.session_state["tab_ct_result"], key_prefix="ct")
